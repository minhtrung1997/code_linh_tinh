# importing the required modules
import glob
import pandas as pd
import os,sys
from openpyxl import load_workbook
import openpyxl

# specifying the path to xlsx files: python merge_xlsx.py <path> <path_containing SNP_analysis file>
path = sys.argv[1]
wkdir = sys.argv[2]

# csv files in the path
file_list = glob.glob(f"{path}/*.xlsx")

###### First way to file merge ##########
# list of excel files we want to merge.
# pd.read_excel(file_path) reads the excel
# data into pandas dataframe.
if False:
    excl_list = []

    for file in file_list:
        excl_list.append(pd.read_excel(file))

    # create a new dataframe to store the
    # merged excel file.
    excl_merged = pd.DataFrame()

    for excl_file in excl_list:
        
        # appends the data into the excl_merged
        # dataframe.
        excl_merged = excl_merged.append(
        excl_file, ignore_index=True)

    # exports the dataframe into excel file with
    # specified name.
    excl_merged.to_excel('total1.xlsx', index=False)

###### Second way to file merge ##########
if True:
    excel_snp = f'{wkdir}/SNP_analysis.xlsx'
    workbook=openpyxl.load_workbook(excel_snp)
    if 'Sheet2' in workbook.sheetnames:
        del workbook['Sheet2']
    if 'Sheet3' in workbook.sheetnames:
        del workbook['Sheet3']
    workbook.save(f'{wkdir}/SNP_analysis.xlsx')
    book = load_workbook(excel_snp)
    writer = pd.ExcelWriter(excel_snp, engine='openpyxl')
    writer.book = book
    for file in file_list:
        id = os.path.basename(file).split('.xlsx')[0]
        df_writefile = pd.read_excel(file)
        df_writefile.to_excel(writer, sheet_name=id, index=None, startcol=0, startrow=0, header=True)
    writer.save()
